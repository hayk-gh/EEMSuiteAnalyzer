from __future__ import annotations
 
import re
import sys
from pathlib import Path
 
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
 
# --------------------------------------------------------------------------- #
# Configuration
# --------------------------------------------------------------------------- #
 
TARGET_YEAR = 2025          # "Focus on 2025"
CUTOFF_DAY = 17              # bills dated after this day roll to the next month
DUP_KEY = ["Facility", "Account Number", "Meter Number", "Service From", "Service To"]
 
RAW_COLUMNS = [
    "Marker", "Facility", "Account Number", "Meter Number", "Provider",
    "Service From", "Service To", "Bill Date", "Consumption", "Demand",
    "Current charges", "Past due", "Total due", "Bill Status",
]
DATA_COLUMNS = [c for c in RAW_COLUMNS if c != "Marker"]
NORMALIZED_COLUMNS = DATA_COLUMNS + ["Source"]
RAW_TAGGED_COLUMNS = ["Source File"] + DATA_COLUMNS
 
# Unit -> CCF (Centum Cubic Feet = 100 cubic feet = 748.052 gallons) conversion factors
WATER_TO_CCF = {
    "ccf": 1.0,                    # already CCF
    "cuft": 1.0 / 100.0,           # cubic feet -> CCF
    "gal": 1.0 / 748.052,          # gallons -> CCF
    "cgal": 100.0 / 748.052,       # "hundred gallons" -> CCF
}
ELECTRIC_UNITS = {"kwh": "kWh", "mwh": "MWh"}
GAS_UNITS = {"therm": "Therms"}
 
FONT = "Arial"
HEADER_FILL = PatternFill("solid", start_color="D9D9D9", end_color="D9D9D9")
HEADER_FONT = Font(name=FONT, bold=True)
 
 
# --------------------------------------------------------------------------- #
# Loading
# --------------------------------------------------------------------------- #
 
def load_raw(path: str | Path) -> pd.DataFrame:
    """Load a bill-export .xlsx: a report-header block in the first few rows,
    then a header row, then data rows, then a trailing 'Total' summary row.
    Works regardless of the file's name."""
    raw = pd.read_excel(path, sheet_name=0, header=None)
 
    header_row_idx = None
    for i in range(min(15, len(raw))):
        if raw.iloc[i].astype(str).str.contains("Facility", na=False).any():
            header_row_idx = i
            break
    if header_row_idx is None:
        raise ValueError(
            f"Could not find a 'Facility' header row in {path}. "
            "Is this a bill export file with the expected layout?"
        )
 
    df = pd.read_excel(path, sheet_name=0, header=None, skiprows=header_row_idx + 2)
    df.columns = RAW_COLUMNS[: df.shape[1]]
    df = df.drop(columns=["Marker"], errors="ignore")
 
    # Drop a trailing summary/"Total" row if present.
    df = df[df["Facility"].notna()].copy()
    df = df[~df["Facility"].astype(str).str.strip().str.lower().eq("total")]
    df = df.reset_index(drop=True)
    return df
 
 
# --------------------------------------------------------------------------- #
# Normalization
# --------------------------------------------------------------------------- #
 
def parse_consumption(value) -> tuple[float | None, str | None]:
    """Split a '97,280.95 kWh' style string into (97280.95, 'kWh')."""
    if pd.isna(value):
        return None, None
    s = str(value).strip()
    m = re.match(r"^([\d,\.\-]+)\s*([A-Za-z]+)\s*$", s)
    if not m:
        return None, None
    num = float(m.group(1).replace(",", ""))
    unit = m.group(2)
    return num, unit
 
 
def compute_bill_period(bill_date: pd.Series) -> pd.Series:
    """Apply the 17th-of-the-month cutoff rule and return a monthly Period."""
    shifted = bill_date.where(bill_date.dt.day <= CUTOFF_DAY, bill_date + pd.DateOffset(months=1))
    return shifted.dt.to_period("M")
 
 
def normalize(df: pd.DataFrame) -> pd.DataFrame:
    """Clean types, fix duplicates (keep 2nd/last entry), and scope to TARGET_YEAR."""
    df = df.copy()
 
    for c in ["Service From", "Service To", "Bill Date"]:
        df[c] = pd.to_datetime(df[c], errors="coerce")
    for c in ["Current charges", "Past due", "Total due"]:
        df[c] = pd.to_numeric(df[c], errors="coerce")
 
    df["Bill Period"] = compute_bill_period(df["Bill Date"])
    df["orig_order"] = range(len(df))
 
    # Fix duplicates: keep the last (2nd) entry per DUP_KEY.
    dup_mask = df.duplicated(subset=DUP_KEY, keep=False)
    keep_idx = set(df.index)
    for _, grp in df[dup_mask].groupby(DUP_KEY):
        grp_sorted = grp.sort_values("orig_order")
        keep_i = grp_sorted.index[-1]
        for drop_i in grp_sorted.index:
            if drop_i != keep_i:
                keep_idx.discard(drop_i)
    df = df.loc[sorted(keep_idx)].copy()
 
    # Scope to target year using the bill-month cutoff rule.
    df = df[df["Bill Period"].apply(lambda p: pd.notna(p) and p.year == TARGET_YEAR)].copy()
 
    df = df.drop(columns=["orig_order", "Bill Period"])
    return df
 
 
# --------------------------------------------------------------------------- #
# Source selection: Approver (single-location) vs. Bill Entry (campus/shared)
# --------------------------------------------------------------------------- #
 
def _match_key(df: pd.DataFrame) -> pd.Series:
    """Facility code + Meter Number + Service dates - the most precise common
    identifier available between the two files."""
    fac_code = df["Facility"].astype(str).str.split("-").str[0].str.strip()
    meter = df["Meter Number"].astype(str).str.strip()
    is_multi = meter.str.lower().eq("multiple") | df["Meter Number"].isna()
    meter = meter.where(~is_multi, "MULTI")
    return (
        fac_code + "|" + meter + "|"
        + df["Service From"].dt.strftime("%Y%m%d") + "|"
        + df["Service To"].dt.strftime("%Y%m%d")
    )
 
 
def select_source(approver_df: pd.DataFrame, bill_entry_df: pd.DataFrame) -> pd.DataFrame:
    """Pick, for each Approver account, whether to use the Approver row or the
    matching Bill Entry row. Temporary rule - see note at top of file."""
    app = approver_df.copy()
    bill = bill_entry_df.copy()
    app["MatchKey"] = _match_key(app)
    bill["MatchKey"] = _match_key(bill)
 
    has_dash = app["Account Number"].astype(str).str.contains("-", na=False)
 
    single_location = app[~has_dash].copy()
    single_location["Source"] = "Approver (single-location)"
 
    campus = app[has_dash].copy()
    campus_keys = campus["MatchKey"].dropna().unique().tolist()
 
    bill_matches = bill[bill["MatchKey"].isin(campus_keys)].copy()
    bill_matches["Source"] = "Bill Entry (campus match)"
    matched_keys = set(bill_matches["MatchKey"])
 
    campus_fallback = campus[~campus["MatchKey"].isin(matched_keys)].copy()
    campus_fallback["Source"] = "Approver (campus - no Bill Entry match)"
 
    combined = pd.concat([single_location, bill_matches, campus_fallback], ignore_index=True)
    return combined.drop(columns=["MatchKey"], errors="ignore")
 
 
# --------------------------------------------------------------------------- #
# Totals by building
# --------------------------------------------------------------------------- #
 
def build_totals_by_building(df: pd.DataFrame) -> pd.DataFrame:
    """Totalize kWh, Therms, and water (converted to CCF) per building (Facility)."""
    d = df.copy()
    d[["Consumption Value", "Consumption Unit"]] = d["Consumption"].apply(
        lambda v: pd.Series(parse_consumption(v))
    )
    d["unit_lower"] = d["Consumption Unit"].str.lower()
 
    rows = []
    for facility, grp in d.groupby("Facility"):
        kwh_total = grp.loc[grp["unit_lower"].isin(ELECTRIC_UNITS), "Consumption Value"].sum()
        therms_total = grp.loc[grp["unit_lower"].isin(GAS_UNITS), "Consumption Value"].sum()
 
        water_ccf_total = 0.0
        water_unit_breakdown = []
        for unit_key, factor in WATER_TO_CCF.items():
            sub = grp.loc[grp["unit_lower"] == unit_key, "Consumption Value"]
            if sub.empty:
                continue
            native_total = sub.sum()
            ccf = native_total * factor
            water_ccf_total += ccf
            water_unit_breakdown.append(f"{round(native_total, 2)} {unit_key}")
 
        rows.append({
            "Facility": facility,
            "Total kWh": round(kwh_total, 2) if kwh_total else 0,
            "Total Therms": round(therms_total, 2) if therms_total else 0,
            "Water Unit(s)": "; ".join(water_unit_breakdown) if water_unit_breakdown else None,
            "Total Water (CCF)": round(water_ccf_total, 2) if water_ccf_total else 0,
        })
 
    result = pd.DataFrame(rows, columns=["Facility", "Total kWh", "Total Therms", "Water Unit(s)", "Total Water (CCF)"])
    return result.sort_values("Facility").reset_index(drop=True)
 
 
# --------------------------------------------------------------------------- #
# Excel output
# --------------------------------------------------------------------------- #
 
def _write_table(ws, df: pd.DataFrame, columns: list[str] | None = None):
    cols = columns or list(df.columns)
    for j, col in enumerate(cols, start=1):
        c = ws.cell(row=1, column=j, value=col)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
    for i, (_, row) in enumerate(df.iterrows(), start=2):
        for j, col in enumerate(cols, start=1):
            val = row.get(col)
            if pd.isna(val):
                val = None
            elif isinstance(val, pd.Timestamp):
                val = val.to_pydatetime()
            elif isinstance(val, pd.Period):
                val = str(val)
            cell = ws.cell(row=i, column=j, value=val)
            cell.font = Font(name=FONT, size=10)
    for j, col in enumerate(cols, start=1):
        if len(df):
            maxlen = df[col].astype(str).str.len().clip(upper=40).max()
            maxlen = 12 if pd.isna(maxlen) else int(maxlen)
        else:
            maxlen = 12
        ws.column_dimensions[get_column_letter(j)].width = max(12, min(40, maxlen + 2))
    ws.freeze_panes = "A2"
 
 
def write_workbook(raw_df: pd.DataFrame, normalized_df: pd.DataFrame, totals_df: pd.DataFrame, output_path: str | Path):
    wb = Workbook()
    wb.remove(wb.active)
 
    ws = wb.create_sheet("Raw Data")
    _write_table(ws, raw_df, columns=RAW_TAGGED_COLUMNS)
 
    ws = wb.create_sheet(f"Normalized {TARGET_YEAR}")
    _write_table(ws, normalized_df, columns=NORMALIZED_COLUMNS)
 
    ws = wb.create_sheet("Totals by Building")
    _write_table(ws, totals_df)
 
    wb.save(output_path)
 
 
# --------------------------------------------------------------------------- #
# Interactive entry point
# --------------------------------------------------------------------------- #
 
def _prompt_for_file(label: str) -> Path:
    while True:
        file_input = input(f"\nEnter the name (or path) of the {label} file: ").strip().strip('"')
        path = Path(file_input)
        if not path.exists():
            if not path.suffix and path.with_suffix(".xlsx").exists():
                path = path.with_suffix(".xlsx")
            else:
                print(f"Could not find a file at: {path}")
                continue
        return path
 
 
def main():
    print("=" * 60)
    print("Bill Normalizer")
    print("=" * 60)
 
    approver_path = _prompt_for_file("Approver")
    bill_entry_path = _prompt_for_file("Bill Entry")
 
    print(f"\nReading {approver_path.name} ...")
    approver_raw = load_raw(approver_path)
    print(f"  {len(approver_raw)} raw rows loaded.")
 
    print(f"Reading {bill_entry_path.name} ...")
    bill_entry_raw = load_raw(bill_entry_path)
    print(f"  {len(bill_entry_raw)} raw rows loaded.")
 
    approver_raw_tagged = approver_raw.copy()
    approver_raw_tagged["Source File"] = approver_path.name
    bill_entry_raw_tagged = bill_entry_raw.copy()
    bill_entry_raw_tagged["Source File"] = bill_entry_path.name
    raw_combined = pd.concat([approver_raw_tagged, bill_entry_raw_tagged], ignore_index=True)
 
    print("Normalizing both files (2025 scope, 17th-of-month cutoff, duplicates fixed) ...")
    approver_norm = normalize(approver_raw)
    bill_entry_norm = normalize(bill_entry_raw)
    print(f"  Approver: {len(approver_norm)} rows in the normalized {TARGET_YEAR} dataset.")
    print(f"  Bill Entry: {len(bill_entry_norm)} rows in the normalized {TARGET_YEAR} dataset.")
 
    print("Selecting source per account (dash in Account Number = campus -> Bill Entry) ...")
    selected_df = select_source(approver_norm, bill_entry_norm)
    print(f"  {(selected_df['Source'] == 'Approver (single-location)').sum()} single-location rows kept from Approver")
    print(f"  {(selected_df['Source'] == 'Bill Entry (campus match)').sum()} campus rows pulled from Bill Entry")
    fallback_n = (selected_df["Source"] == "Approver (campus - no Bill Entry match)").sum()
    print(f"  {fallback_n} campus rows had no Bill Entry match - kept Approver as fallback (flagged, worth a spot check)")
 
    print("Totalizing by building ...")
    totals_df = build_totals_by_building(selected_df)
 
    output_path = Path(f"{TARGET_YEAR}_TotalGHGEmissions.xlsx")
    write_workbook(raw_combined, selected_df, totals_df, output_path)
    print(f"\nSaved: {output_path.resolve()}")
 
 
if __name__ == "__main__":
    main()